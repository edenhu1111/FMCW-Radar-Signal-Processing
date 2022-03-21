import numpy as np
from numpy.core.function_base import linspace
from scipy.fftpack import fft,ifft
import matplotlib.pyplot as plt
from czt import czt
import openpyxl
from os_cfar import os_cfar

empty_data = openpyxl.load_workbook("20210114_empty000.xlsx")
data_2_5m = openpyxl.load_workbook("20210114_2_5M000.xlsx")
data_4_m = openpyxl.load_workbook("20210114_4M000.xlsx")

ws0 = empty_data['20210114_empty000']
ws1 = data_2_5m['20210114_2_5M000']
ws2 = data_4_m['20210114_4M000']
dataEmpty = np.zeros(10000)
data2_5M = np.zeros(10000)
data4M = np.zeros(10000)

for ii in range(1,10001):
    dataEmpty[ii-1] = ws0.cell(row = ii,column = 5).value
    data2_5M[ii-1]  = ws1.cell(row = ii,column = 5).value
    data4M[ii-1]    = ws2.cell(row = ii,column = 5).value

w = np.hamming(10000)
pulseCompress1 = fft(w*dataEmpty)
pulseCompress2 = fft(w*data2_5M)
pulseCompress3 = fft(w*data4M)

freq_x = linspace(0,1/2.00E-08,10000,endpoint=False)

f1 = plt.subplot(1,3,1)
plt.plot(freq_x[0:1500],abs(pulseCompress1[0:1500]))
f1.set_title("0m")
plt.xlabel('frequency')
plt.ylim([0,40])

f2 = plt.subplot(1,3,2)
plt.plot(freq_x[0:1500],abs(pulseCompress2[0:1500]))
f2.set_title("2.5m")
plt.xlabel('frequency')
plt.ylim([0,40])

f3 = plt.subplot(1,3,3)
plt.plot(freq_x[0:1500],abs(pulseCompress3[0:1500]))
f3.set_title("4m")
plt.xlabel('frequency')
plt.ylim([0,40])


plt.show()